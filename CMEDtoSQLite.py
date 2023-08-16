## INICIO DO SCRIPT ##
import sqlite3
import argparse
import time
from openpyxl import load_workbook
from tqdm import tqdm

# Configurar argumentos da linha de comando
parser = argparse.ArgumentParser(description="Converter planilha CMED da ANVISA para Banco de Dados SQLite3.")
parser.add_argument('-sheet', type=str, required=True, help='Caminho para o arquivo da planilha Excel')
parser.add_argument('-db', type=str, help='Caminho para o arquivo do banco de dados SQLite')
args = parser.parse_args()

# Obter o caminho do arquivo Excel a partir dos argumentos da linha de comando
excel_file = args.sheet
db_file = args.db

# Verificar se o caminho do arquivo Excel foi fornecido
if not excel_file:
    print("Erro: Caminho do arquivo Excel não fornecido.")
    parser.print_help()
    exit(1)

if not db_file:
    print("Criando banco de dados...")
    
    # Create the database file
    db_file = 'LISTACMED.db'
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS "SUBSTANCIAS" (
            "SubstanciaID"	INTEGER NOT NULL UNIQUE,
            "Substancia"	TEXT,
            PRIMARY KEY("SubstanciaID" AUTOINCREMENT)
        );
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS "LABORATORIOS" (
            "LaboratorioID" INTEGER NOT NULL UNIQUE,
            "CNPJ" TEXT,
            "Laboratorio" TEXT,
            PRIMARY KEY("LaboratorioID" AUTOINCREMENT)
        );
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS "CLASSES_TERAPEUTICAS" (
            "ClasseTerapeuticaID" INTEGER NOT NULL UNIQUE,
            "CodigoClasse" TEXT,
            "DescricaoClasse" TEXT,
            PRIMARY KEY("ClasseTerapeuticaID" AUTOINCREMENT)
        );
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS "PRODUTOS" (
            "ProdutoID"	INTEGER NOT NULL UNIQUE,
            "SubstanciaID"	INTEGER,
            "LaboratorioID"	INTEGER,
            "ClasseTerapeuticaID"	INTEGER,
            "GGREM"	INTEGER,
            "Registro"	INTEGER,
            "EAN1"	INTEGER,
            "EAN2"	INTEGER,
            "EAN3"	INTEGER,
            "Produto"	TEXT,
            "Apresentacao"	TEXT,
            "Tipo"	INTEGER,
            "RegimePreco"	INTEGER,
            "RestricaoHospitalar"	INTEGER,
            "CAP"	INTEGER,
            "CONFAZ87"	INTEGER,
            "ICMS_0"	INTEGER,
            "Lista"	INTEGER,
            "ComercializaAnoAnterior"	INTEGER,
            "Tarja"	INTEGER,
            FOREIGN KEY("ClasseTerapeuticaID") REFERENCES "CLASSES_TERAPEUTICAS"("ClasseTerapeuticaID") ON UPDATE CASCADE ON DELETE RESTRICT,
            FOREIGN KEY("LaboratorioID") REFERENCES "LABORATORIOS"("LaboratorioID") ON UPDATE CASCADE ON DELETE RESTRICT,
            FOREIGN KEY("SubstanciaID") REFERENCES "SUBSTANCIAS"("SubstanciaID") ON UPDATE CASCADE ON DELETE RESTRICT,
            PRIMARY KEY("ProdutoID" AUTOINCREMENT)
        );
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS "PRECOS" (
            "PrecoID" INTEGER NOT NULL UNIQUE,
            "ProdutoID" INTEGER NOT NULL,
            "PFSemImposto" NUMERIC,
            "PF_0" NUMERIC,
            "PF_12" NUMERIC,
            "PF_17" NUMERIC,
            "PF_17_ALC" NUMERIC,
            "PF_17_5" NUMERIC,
            "PF_17_5_ALC" NUMERIC,
            "PF_18" NUMERIC,
            "PF_18_ALC" NUMERIC,
            "PF_19" NUMERIC,
            "PF_20" NUMERIC,
            "PF_21" NUMERIC,
            "PF_22" NUMERIC,
            "PMC_0" NUMERIC,
            "PMC_12" NUMERIC,
            "PMC_17" NUMERIC,
            "PMC_17_ALC" NUMERIC,
            "PMC_17_5" NUMERIC,
            "PMC_17_5_ALC" NUMERIC,
            "PMC_18" NUMERIC,
            "PMC_18_ALC" NUMERIC,
            "PMC_19" NUMERIC,
            "PMC_20" NUMERIC,
            "PMC_21" NUMERIC,
            "PMC_22" NUMERIC,
            FOREIGN KEY("ProdutoID") REFERENCES "PRODUTOS"("ProdutoID") ON UPDATE CASCADE ON DELETE RESTRICT,
            PRIMARY KEY("PrecoID" AUTOINCREMENT)
        );
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS "criacao_db" (
	        "DATA"	INTEGER
        );
    ''')

    conn.commit()
    conn.close()

sinonimos_substancia = {
    "DIPIRONA MONOIDRATADA": "DIPIRONA",
    "AMOXICILINA TRIHIDRATADA": "AMOXICILINA",
    "AMOXICILINA TRI-HIDRATADA": "AMOXICILINA",
    "MALEATO DE CLORFENAMINA": "MALEATO DE CLORFENIRAMINA",
    "DEXCLORFENIRAMINA": "MALEATO DE DEXCLORFENIRAMINA",
    "CAFEÍNA ANIDRA": "CAFEÍNA"
}

# Função para normalizar as substancias
def normaliza_substancia(substancia_str):
    listaSubstancia = substancia_str.split(';')

    substanciasNormalizada = []
    for subs in listaSubstancia:
        substanciaNormalizada = sinonimos_substancia.get(subs, subs)
        substanciasNormalizada.append(substanciaNormalizada)
    
    substanciasNormalizada = sorted(substanciasNormalizada)
    return ';'.join(substanciasNormalizada)

# Função para mapear valores 'Sim' e 'Não' para 1 e 0
def map_sim_nao(value):
    return 1 if value == 'Sim' else 0

# Função para retornar nulo quando houve o caractere '-' na tabela
def map_dash_none(value):
    return None if value == '-' else value

# Função para tratar os preços
def price_to_float(value):
    try:
        return float(value.replace(",","."))
    except (ValueError, AttributeError):
        return None

# Função para mapear valores 'Positiva', 'Negativa' e 'Neutra' para 1, 2 e 3
def map_pos_neg_neu(value):
    if value == 'Positiva':
        return 1
    elif value == 'Negativa':
        return 2
    elif value == 'Neutra':
        return 3
    return None

# Função para mapear o tipo de medicamento
def map_tipo(value):
    if value == 'Genérico':
        return 1
    elif value == 'Similar':
        return 2
    elif value == 'Novo':
        return 3
    elif value == 'Biológico':
        return 4
    elif value == 'Específico':
        return 5
    elif value == 'Fitoterápico':
        return 6
    elif value == 'Produto de Terapia Avançada':
        return 7
    elif value == 'Radiofármaco':
        return 8
    elif value == 'Regulado':
        return 9
    return None
    
# Função para mapear os valoras dos tipos de tarja
def map_tarjado(value):
    if value == 'Tarja Sem Tarja':
        return 1
    elif value == 'Tarja Preta' or value == 'Tarja Preta (**)':
        return 2
    elif value == 'Tarja Vermelha' or value == 'Tarja Vermelha (**)' or value == 'Tarja Vermelha sob restrição':
        return 3
    return None

# Função para mapear valores 'Liberado' e 'Regulado' para 1 e 2
def map_regime_preco(value):
    return 1 if value == 'Liberado' else 2

# Carregar o arquivo Excel
print("Carregando planilha...")
print("O tempo de carragamento depende da quantidade de registros...")
workbook = load_workbook(excel_file)
sheet = workbook.active
print("Carregamento concluído!")

# Conectar ao banco de dados SQLite
conn = sqlite3.connect(db_file)
cursor = conn.cursor()

# PROCESSA AS SUBSTANCIAS
for row in tqdm(sheet.iter_rows(min_row=2, values_only=True), desc="Processando SUBSTANCIAS"):
    substancia = row[0]  # A primeira coluna contém o nome da substância

    # Verificar se a substância já existe no banco de dados
    cursor.execute("SELECT SubstanciaID FROM SUBSTANCIAS WHERE Substancia = ?", (substancia,))
    existing_substance = cursor.fetchone()

    if not existing_substance:
        # Inserir a substância no banco de dados
        cursor.execute("INSERT INTO SUBSTANCIAS (Substancia) VALUES (?)", (substancia,))
        conn.commit()

# PROCESSA OS LABORATÓRIOS
for row in tqdm(sheet.iter_rows(min_row=2, values_only=True), desc="Processando LABORATORIOS"):
    cnpj = row[1]  # Coluna 2 contém o CNPJ
    laboratorio = row[2]  # Coluna 3 contém o Laboratório

    # Verificar se o CNPJ já existe no banco de dados
    cursor.execute("SELECT LaboratorioID FROM LABORATORIOS WHERE CNPJ = ?", (cnpj,))
    existing_lab = cursor.fetchone()

    if not existing_lab:
        cursor.execute("INSERT INTO LABORATORIOS (CNPJ, Laboratorio) VALUES (?, ?)", (cnpj, laboratorio))
        conn.commit()

# PROCESSA AS CLASSES TERAPEUTICAS
for index, row in enumerate(tqdm(sheet.iter_rows(min_row=2, values_only=True), desc="Processando CLASSES TERAPEUTICAS"), start=2):
    try:
        classe_terapeutica = row[10]  # Décima coluna contém o valor a ser dividido

        # Dividir o valor usando o caractere '-' e remover espaços em branco
        sub_string = classe_terapeutica.split('-')

        if len(sub_string) >= 2:
            codigo_classe = sub_string[0].strip()
            descricao_classe = '-'.join(sub_string[1:]).strip()

            # Verificar se o CodigoClasse já existe no banco de dados
            cursor.execute("SELECT ClasseTerapeuticaID FROM CLASSES_TERAPEUTICAS WHERE CodigoClasse = ?", (codigo_classe,))
            existing_class = cursor.fetchone()

            if not existing_class:
                cursor.execute("INSERT INTO CLASSES_TERAPEUTICAS (CodigoClasse, DescricaoClasse) VALUES (?, ?)", (codigo_classe, descricao_classe))
                conn.commit()

    except Exception as e:
        print(f"Erro na linha {index}: {e}")


# PROCESSA OS PRODUTOS
for row in tqdm(sheet.iter_rows(min_row=2, values_only=True), desc="Processando PRODUTOS"):
    ggrem = row[3]
    registro = row[4]
    ean1 = row[5]
    ean2 = map_dash_none(row[6].strip())
    ean3 = map_dash_none(row[7].strip())
    produto = row[8]
    apresentacao = row[9]
    tipo = map_tipo(row[11])
    regime_preco = map_regime_preco(row[12])
    restricao_hospitalar = map_sim_nao(row[38])
    cap = map_sim_nao(row[39])
    confaz87 = map_sim_nao(row[40])
    icms_0 = map_sim_nao(row[41])
    lista = map_pos_neg_neu(row[43])
    comercializa_ano_anterior = map_sim_nao(row[44])
    tarja = map_tarjado(row[45])

    # Verificar se o EAN1 já existe no banco de dados
    cursor.execute("SELECT ProdutoID FROM PRODUTOS WHERE EAN1 = ?", (ean1,))
    existing_product = cursor.fetchone()

    if not existing_product:
        # Obter LaboratorioID a partir do CNPJ
        cnpj = row[1]
        cursor.execute("SELECT LaboratorioID FROM LABORATORIOS WHERE CNPJ = ?", (cnpj,))
        laboratorio_id = cursor.fetchone()

        if laboratorio_id:
            laboratorio_id = laboratorio_id[0]
        else:
            laboratorio_id = None

        # Obter SubstanciaID a partir do Substance
        substancia = row[0]
        cursor.execute("SELECT SubstanciaID FROM SUBSTANCIAS WHERE Substancia = ?", (substancia,))
        substancia_id = cursor.fetchone()

        if substancia_id:
            substancia_id = substancia_id[0]
        else:
            substancia_id = None

        # Obter ClasseTerapeuticaID a partir da coluna dividida
        codigo_classe = row[10].split('-')[0].strip()
        cursor.execute("SELECT ClasseTerapeuticaID FROM CLASSES_TERAPEUTICAS WHERE CodigoClasse = ?", (codigo_classe,))
        classe_terapeutica_id = cursor.fetchone()

        if classe_terapeutica_id:
            classe_terapeutica_id = classe_terapeutica_id[0]
        else:
            classe_terapeutica_id = None

        # Inserir os dados no banco de dados
        cursor.execute("""
            INSERT INTO PRODUTOS (
                SubstanciaID, LaboratorioID, ClasseTerapeuticaID, GGREM, Registro, EAN1, EAN2, EAN3,
                Produto, Apresentacao, Tipo, RegimePreco, RestricaoHospitalar, CAP, CONFAZ87, ICMS_0,
                Lista, ComercializaAnoAnterior, Tarja
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            substancia_id, laboratorio_id, classe_terapeutica_id, ggrem, registro, ean1, ean2, ean3,
            produto, apresentacao, tipo, regime_preco, restricao_hospitalar, cap, confaz87, icms_0,
            lista, comercializa_ano_anterior, tarja
        ))

        conn.commit()

# PROCESSA OS PREÇOS
for row in tqdm(sheet.iter_rows(min_row=2, values_only=True), desc="Processando PRECOS"):
    pf_tax_free = price_to_float(row[13])
    pf_0 = price_to_float(row[14])
    pf_12 = price_to_float(row[15])
    pf_17 = price_to_float(row[16])
    pf_17_alc = price_to_float(row[17])
    pf_17_5 = price_to_float(row[18])
    pf_17_5_alc = price_to_float(row[19])
    pf_18 = price_to_float(row[20])
    pf_18_alc = price_to_float(row[21])
    pf_19 = price_to_float(row[22])
    pf_20 = price_to_float(row[23])
    pf_21 = price_to_float(row[24])
    pf_22 = price_to_float(row[25])
    pmc_0 = price_to_float(row[26])
    pmc_12 = price_to_float(row[27])
    pmc_17 = price_to_float(row[28])
    pmc_17_alc = price_to_float(row[29])
    pmc_17_5 = price_to_float(row[30])
    pmc_17_5_alc = price_to_float(row[31])
    pmc_18 = price_to_float(row[32])
    pmc_18_alc = price_to_float(row[33])
    pmc_19 = price_to_float(row[34])
    pmc_20 = price_to_float(row[35])
    pmc_21 = price_to_float(row[36])
    pmc_22 = price_to_float(row[37])

    # Pega a ID do Produto usando EAN13
    ean1 = row[5]
    cursor.execute("SELECT ProdutoID FROM PRODUTOS WHERE EAN1 = ?", (ean1,))
    product_id = cursor.fetchone()

    if product_id:
        product_id = product_id[0]
    else:
        product_id = None

    # Verifica se o preço do produto já foi cadastrado antes
    cursor.execute("SELECT PrecoID FROM PRECOS WHERE ProdutoID = ?", (product_id,))
    existing_price = cursor.fetchone()

    if not existing_price:

        # Inserir os dados no banco de dados
        cursor.execute("""
            INSERT INTO PRECOS (
                ProdutoID, PFSemImposto, PF_0, PF_12, PF_17, PF_17_ALC, PF_17_5, PF_17_5_ALC,
                PF_18, PF_18_ALC, PF_19, PF_20, PF_21, PF_22, PMC_0, PMC_12, PMC_17, PMC_17_ALC,
                PMC_17_5, PMC_17_5_ALC, PMC_18, PMC_18_ALC, PMC_19, PMC_20, PMC_21, PMC_22
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            product_id, pf_tax_free, pf_0, pf_12, pf_17, pf_17_alc, pf_17_5, pf_17_5_alc,
            pf_18, pf_18_alc, pf_19, pf_20, pf_21, pf_22, pmc_0, pmc_12, pmc_17, pmc_17_alc,
            pmc_17_5, pmc_17_5_alc, pmc_18, pmc_18_alc, pmc_19, pmc_20, pmc_21, pmc_22
        ))

        conn.commit()

# Salva a data de criação do banco
timestamp = time.time_ns()
cursor.execute("INSERT INTO criacao_db (DATA) VALUES (?)", (timestamp,))
conn.commit()

# Fechar a conexão com o banco de dados
conn.close()

print("Conversão finalizada.")
## FIM DO SCRIPT ##
